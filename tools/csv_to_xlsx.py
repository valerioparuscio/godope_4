"""Convert a UTF-8 CSV (game data draft) into a formatted .xlsx workbook.

Usage:
    python tools/csv_to_xlsx.py <input.csv> [output.xlsx]

If output.xlsx is omitted, it is written next to the input file with the
same name and an .xlsx extension.
"""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_WIDTH = 12
MAX_WIDTH = 60


def convert(src: Path, dst: Path) -> int:
    with src.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))

    wb = Workbook()
    ws = wb.active
    ws.title = src.stem[:31]

    header = rows[0]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows[1:]:
        ws.append(row)

    for i, name in enumerate(header, start=1):
        content_width = max((len(r[i - 1]) for r in rows if i - 1 < len(r)), default=0)
        width = min(max(len(name), content_width // 2, DEFAULT_WIDTH), MAX_WIDTH)
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(dst)
    return len(rows) - 1


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    src_path = Path(sys.argv[1])
    dst_path = Path(sys.argv[2]) if len(sys.argv) > 2 else src_path.with_suffix(".xlsx")

    n = convert(src_path, dst_path)
    print(f"Wrote {dst_path} ({n} data rows)")
